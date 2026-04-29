import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ─── NOMBRE DE ARCHIVO CON FECHA ────────────────────────────────────────────
OUTPUT_FILE = f"BBR_Propiedades_Estandar_{date.today().strftime('%Y%m%d')}.xlsx"

wb = openpyxl.Workbook()

# ─── COLORES ────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1E3A5F"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_GROUP_BG    = "D6E4F0"
COLOR_GROUP_FONT  = "1E3A5F"
COLOR_ROW_EVEN    = "F5F9FC"
COLOR_ROW_ODD     = "FFFFFF"
COLOR_REQUIRED    = "FFF3CD"
COLOR_BORDER      = "B0C4D8"

thin   = Side(style="thin", color=COLOR_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── COLUMNAS ────────────────────────────────────────────────────────────────
# (header, ancho, obligatorio, grupo)
COLUMNS = [
    # IDENTIFICACIÓN
    ("ID",                   12, True,  "IDENTIFICACIÓN"),
    ("Tipo",                 16, True,  "IDENTIFICACIÓN"),
    ("Operación",            14, True,  "IDENTIFICACIÓN"),
    ("Activo",               10, True,  "IDENTIFICACIÓN"),
    ("Destacado",            12, False, "IDENTIFICACIÓN"),

    # UBICACIÓN
    ("Título",               30, True,  "UBICACIÓN"),
    ("Calle",                22, True,  "UBICACIÓN"),
    ("Barrio",               18, False, "UBICACIÓN"),
    ("Ciudad",               16, True,  "UBICACIÓN"),

    # PRECIO
    ("Precio",               14, False, "PRECIO"),
    ("Moneda",               10, True,  "PRECIO"),
    ("Expensas",             12, False, "PRECIO"),

    # CARACTERÍSTICAS
    ("Antigüedad",           14, False, "CARACTERÍSTICAS"),
    ("Estado Construcción",  20, False, "CARACTERÍSTICAS"),
    ("Ambientes",            12, False, "CARACTERÍSTICAS"),
    ("Dormitorios",          14, False, "CARACTERÍSTICAS"),
    ("Baños",                10, False, "CARACTERÍSTICAS"),
    ("Sup. Total",           14, False, "CARACTERÍSTICAS"),
    ("Sup. Cubierta",        14, False, "CARACTERÍSTICAS"),

    # DESCRIPCIONES
    ("Descripción Larga",    60, True,  "DESCRIPCIONES"),
    ("Descripción Corta",    40, True,  "DESCRIPCIONES"),

    # FOTOS
    ("Carpeta",              10, True,  "FOTOS"),
    ("URL Foto 1",           50, True,  "FOTOS"),
    ("URL Foto 2",           50, False, "FOTOS"),
    ("URL Foto 3",           50, False, "FOTOS"),
    ("URL Foto 4",           50, False, "FOTOS"),
    ("URL Foto 5",           50, False, "FOTOS"),

    # DETALLES
    ("Detalles",             45, False, "DETALLES"),
]

# ════════════════════════════════════════════════════════════════════════════
#  HOJA 1: Propiedades
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Propiedades"
ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 22

group_fill  = PatternFill("solid", fgColor=COLOR_GROUP_BG)
group_font  = Font(bold=True, color=COLOR_GROUP_FONT, size=9)
group_align = Alignment(horizontal="center", vertical="center")

header_fill  = PatternFill("solid", fgColor=COLOR_HEADER_BG)
header_font  = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
req_fill     = PatternFill("solid", fgColor=COLOR_REQUIRED)

prev_group  = None
group_start = 1

for col_idx, (header, width, required, group) in enumerate(COLUMNS, start=1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

    # Fila 2: header
    cell2 = ws.cell(row=2, column=col_idx, value=header)
    cell2.fill      = req_fill if required else header_fill
    cell2.font      = Font(bold=True, color="1E3A5F" if required else COLOR_HEADER_FONT, size=10)
    cell2.alignment = header_align
    cell2.border    = border

    # Fila 1: grupo
    cell1 = ws.cell(row=1, column=col_idx, value=group if group != prev_group else "")
    cell1.fill      = group_fill
    cell1.font      = group_font
    cell1.alignment = group_align
    cell1.border    = border

    if group != prev_group:
        if prev_group is not None and col_idx - 1 > group_start:
            ws.merge_cells(start_row=1, start_column=group_start,
                           end_row=1,   end_column=col_idx - 1)
        group_start = col_idx
        prev_group  = group

# merge último grupo
ws.merge_cells(start_row=1, start_column=group_start,
               end_row=1,   end_column=len(COLUMNS))

# ─── FILAS DE EJEMPLO ────────────────────────────────────────────────────────
# Cada fila tiene exactamente len(COLUMNS) = 29 valores
SAMPLE_ROWS = [
    [
        "PROP-001", "casa", "venta", "si", "si",
        "Casa Venta - Ramallo", "Colón al 1100", "Centro", "Ramallo",
        "", "USD", "",
        "", "usado", "", 2, 1, "462 m²", "140 m²",
        "Propiedad de aprox. 140 m2, en lote de 11 m. x 42 m. sobre calle Colón al 1100...",
        "Excelente ubicación en zona céntrica y comercial de Ramallo.",
        "1",
        "https://res.cloudinary.com/dikb9wzup/image/upload/v1766830381/bbr/prop-001/foto01.jpg",
        "https://res.cloudinary.com/dikb9wzup/image/upload/v1766830382/bbr/prop-001/foto02.jpg",
        "https://res.cloudinary.com/dikb9wzup/image/upload/v1766830383/bbr/prop-001/foto03.jpg",
        "", "",
        "patio, pileta, quincho, parrilla",
    ],
    [
        "PROP-002", "departamento", "alquiler", "si", "no",
        "Dpto Alquiler - Ramallo", "San Martín al 400", "Centro", "Ramallo",
        "150000", "ARS", "8000",
        "5 años", "usado", 2, 1, 1, "55 m²", "55 m²",
        "Departamento 1 dormitorio totalmente equipado, piso 3ro con vista a la plaza...",
        "Dpto moderno en pleno centro, todos los servicios incluidos.",
        "2",
        "https://res.cloudinary.com/.../foto01.jpg",
        "", "", "", "",
        "cochera, balcon",
    ],
]

data_align   = Alignment(vertical="center", wrap_text=False)
data_align_w = Alignment(vertical="center", wrap_text=True)
WRAP_COLS    = {20, 21, 23, 24, 25, 26, 27}  # descripciones y URLs

for row_idx, row_data in enumerate(SAMPLE_ROWS, start=3):
    ws.row_dimensions[row_idx].height = 40
    fill_color = COLOR_ROW_EVEN if row_idx % 2 == 0 else COLOR_ROW_ODD
    row_fill   = PatternFill("solid", fgColor=fill_color)
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill      = row_fill
        cell.border    = border
        cell.alignment = data_align_w if col_idx in WRAP_COLS else data_align
        cell.font      = Font(size=10)

# ─── FILA DE REFERENCIA ───────────────────────────────────────────────────────
ref_row = len(SAMPLE_ROWS) + 4
ws.row_dimensions[ref_row].height = 60
ref_data = [
    "PROP-XXX",
    "casa / departamento /\nlote / campo",
    "venta / alquiler",
    "si / no",
    "si / no",
    "Texto libre",
    "Texto libre",
    "Texto libre",
    "Texto libre",
    "Número (0 = sin precio)",
    "USD / ARS",
    "Número (0 = sin expensas)",
    "Ej: 10 años / 2 años / sin dato",
    "nuevo / usado / en construccion",
    "Número",
    "Número",
    "Número",
    "Ej: 120 m²",
    "Ej: 90 m²",
    "Texto largo libre",
    "Texto corto (max ~150 car.)",
    "Número de carpeta",
    "URL completa",
    "URL completa",
    "URL completa",
    "URL completa",
    "URL completa",
    "Separados por coma:\npileta, cochera, parrilla...",
]

ref_fill  = PatternFill("solid", fgColor="E8F0E9")
ref_font  = Font(italic=True, color="2D6A2D", size=9)
ref_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

label_cell = ws.cell(row=ref_row - 1, column=1, value="REFERENCIA DE VALORES PERMITIDOS")
label_cell.font      = Font(bold=True, color="2D6A2D", size=10)
label_cell.alignment = Alignment(vertical="center")
ws.merge_cells(start_row=ref_row - 1, start_column=1,
               end_row=ref_row - 1,   end_column=len(COLUMNS))

for col_idx, ref_value in enumerate(ref_data, start=1):
    cell = ws.cell(row=ref_row, column=col_idx, value=ref_value)
    cell.fill      = ref_fill
    cell.font      = ref_font
    cell.alignment = ref_align
    cell.border    = border

ws.freeze_panes   = "A3"
ws.sheet_view.zoomScale = 90

# ════════════════════════════════════════════════════════════════════════════
#  HOJA 2: Instrucciones
# ════════════════════════════════════════════════════════════════════════════
wi = wb.create_sheet("Instrucciones")
wi.sheet_view.zoomScale = 100
wi.column_dimensions["A"].width = 22
wi.column_dimensions["B"].width = 20
wi.column_dimensions["C"].width = 14
wi.column_dimensions["D"].width = 55

title_font   = Font(bold=True, color="FFFFFF", size=13)
title_fill   = PatternFill("solid", fgColor="1E3A5F")
title_align  = Alignment(horizontal="center", vertical="center")

sec_font  = Font(bold=True, color="1E3A5F", size=11)
sec_fill  = PatternFill("solid", fgColor="D6E4F0")
sec_align = Alignment(vertical="center")

field_font_bold  = Font(bold=True, size=10)
field_font       = Font(size=10)
field_align      = Alignment(vertical="top", wrap_text=True)
oblig_font       = Font(size=10, color="8B6A00")
oblig_fill       = PatternFill("solid", fgColor="FFF3CD")

note_font  = Font(italic=True, size=9, color="555555")
note_fill  = PatternFill("solid", fgColor="F0F0F0")
note_align = Alignment(vertical="top", wrap_text=True)

thin2   = Side(style="thin", color="BBBBBB")
border2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

def write_title(ws, row, text):
    ws.row_dimensions[row].height = 28
    c = ws.cell(row=row, column=1, value=text)
    c.font = title_font; c.fill = title_fill; c.alignment = title_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

def write_section(ws, row, text):
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=1, value=text)
    c.font = sec_font; c.fill = sec_fill; c.alignment = sec_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

def write_header_row(ws, row):
    ws.row_dimensions[row].height = 18
    for col, val in enumerate(["Campo", "Grupo JSON", "Obligatorio", "Descripción / Valores permitidos"], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border2

def write_field(ws, row, campo, grupo, obligatorio, descripcion):
    ws.row_dimensions[row].height = 36
    is_oblig = obligatorio == "SI"
    for col, val in enumerate([campo, grupo, obligatorio, descripcion], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border2
        c.alignment = field_align
        if col == 3 and is_oblig:
            c.font = oblig_font; c.fill = oblig_fill
        elif col == 1:
            c.font = field_font_bold
        else:
            c.font = field_font

def write_note(ws, row, text):
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value=text)
    c.font = note_font; c.fill = note_fill; c.alignment = note_align
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

r = 1
write_title(wi, r, f"BBR Grupo Inmobiliario — Instrucciones de carga de propiedades  |  v{date.today().strftime('%Y%m%d')}")

r += 2
write_section(wi, r, "CONSIDERACIONES GENERALES")
r += 1
notes = [
    "• Completar UNA FILA por propiedad en la hoja 'Propiedades'.",
    "• Los campos en AMARILLO son OBLIGATORIOS. El resto es opcional pero recomendado.",
    "• Los valores nulos (sin dato) deben dejarse en blanco (no escribir 'null' ni 'N/A').",
    "• Las URLs de fotos deben ser públicas y accesibles (recomendado: Cloudinary).",
    "• El archivo actualizado se sube al repositorio GitHub para que la landing se actualice automáticamente.",
]
for note in notes:
    r += 1
    wi.row_dimensions[r].height = 18
    c = wi.cell(row=r, column=1, value=note)
    c.font = Font(size=10); c.alignment = Alignment(vertical="center", wrap_text=True)
    wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

r += 2
write_section(wi, r, "DETALLE DE CAMPOS")
r += 1
write_header_row(wi, r)

fields = [
    # (campo, grupo JSON, obligatorio, descripción)
    ("ID",                  "raíz",             "SI",  "Identificador único. Formato: PROP-001, PROP-002, etc. No repetir."),
    ("Tipo",                "raíz",             "SI",  "Tipo de propiedad. Valores: casa | departamento | lote | campo"),
    ("Operación",           "raíz",             "SI",  "Tipo de operación. Valores: venta | alquiler"),
    ("Activo",              "raíz",             "SI",  "Si la propiedad está activa y visible en la web. Valores: si | no"),
    ("Destacado",           "raíz",             "NO",  "Si aparece destacada en la sección principal. Valores: si | no. Default: no"),
    ("Título",              "raíz",             "SI",  "Título descriptivo. Ej: 'Casa Venta - Ramallo' o 'Dpto Alquiler - Centro'."),
    ("Calle",               "direccion",        "SI",  "Calle y altura. Ej: 'Colón al 1100' o 'San Martín 450'."),
    ("Barrio",              "direccion",        "NO",  "Barrio o zona. Ej: 'Centro', 'Las Acacias'. Dejar vacío si no aplica."),
    ("Ciudad",              "direccion",        "SI",  "Ciudad. Ej: 'Ramallo', 'San Pedro'."),
    ("Precio",              "precio",           "NO",  "Valor numérico sin separadores. Ej: 75000. Dejar en 0 o vacío si es 'Consultar'."),
    ("Moneda",              "precio",           "SI",  "Moneda del precio. Valores: USD | ARS"),
    ("Expensas",            "precio",           "NO",  "Monto mensual de expensas en ARS. Solo aplica a departamentos. Dejar vacío si no tiene."),
    ("Antigüedad",          "caracteristicas",  "NO",  "Antigüedad de la propiedad. Ej: '10 años', '2 años', 'a estrenar'. Dejar vacío si no se sabe."),
    ("Estado Construcción", "caracteristicas",  "NO",  "Estado de la construcción. Valores: nuevo | usado | en construccion"),
    ("Ambientes",           "caracteristicas",  "NO",  "Cantidad total de ambientes (número). Dejar vacío si no aplica."),
    ("Dormitorios",         "caracteristicas",  "NO",  "Cantidad de dormitorios (número). Dejar vacío si no aplica."),
    ("Baños",               "caracteristicas",  "NO",  "Cantidad de baños (número). Dejar vacío si no aplica."),
    ("Sup. Total",          "caracteristicas",  "NO",  "Superficie total del terreno. Incluir unidad. Ej: '462 m²'."),
    ("Sup. Cubierta",       "caracteristicas",  "NO",  "Superficie cubierta edificada. Incluir unidad. Ej: '140 m²'."),
    ("Descripción Larga",   "raíz",             "SI",  "Descripción completa de la propiedad. Sin límite de extensión. Usar lenguaje claro y detallado."),
    ("Descripción Corta",   "raíz",             "SI",  "Resumen de 1-2 oraciones para mostrar en cards. Máximo ~150 caracteres."),
    ("Carpeta",             "fotos",            "SI",  "Número de carpeta para organizar fotos en Cloudinary. Ej: 1, 2, 3..."),
    ("URL Foto 1",          "fotos",            "SI",  "URL pública de la foto principal. Formato Cloudinary recomendado."),
    ("URL Foto 2",          "fotos",            "NO",  "URL pública de foto adicional. Dejar vacío si no hay."),
    ("URL Foto 3",          "fotos",            "NO",  "URL pública de foto adicional. Dejar vacío si no hay."),
    ("URL Foto 4",          "fotos",            "NO",  "URL pública de foto adicional. Dejar vacío si no hay."),
    ("URL Foto 5",          "fotos",            "NO",  "URL pública de foto adicional. Dejar vacío si no hay."),
    ("Detalles",            "raíz",             "NO",  "Amenities y características extra, separados por coma. Ej: pileta, cochera, parrilla, quincho, jardín, patio."),
]

for campo, grupo, oblig, desc in fields:
    r += 1
    write_field(wi, r, campo, grupo, oblig, desc)

r += 2
write_section(wi, r, "NOTAS DE ACTUALIZACIÓN")
r += 1
changelog = [
    f"v{date.today().strftime('%Y%m%d')} — Campos nuevos: Expensas (sección PRECIO), Antigüedad y Estado Construcción (sección CARACTERÍSTICAS, antes en raíz), Activo y Destacado (sección IDENTIFICACIÓN).",
    "v20250101 (v3) — Incorporación de campos Activa y Título al template.",
    "v20240901 (v2) — Template inicial BBR con estructura de propiedades estándar.",
]
for entry in changelog:
    r += 1
    wi.row_dimensions[r].height = 22
    c = wi.cell(row=r, column=1, value=entry)
    c.font = Font(size=9, italic=True, color="444444")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

wi.freeze_panes = "A1"

# ─── GUARDAR ────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"Archivo generado: {OUTPUT_FILE}")
