#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CREAR EXCEL TEMPLATE PARA INMOBILIARIAS
========================================
Script para generar el archivo Excel plantilla que las inmobiliarias deben completar
con sus propiedades. El Excel resultante tiene validaciones, listas desplegables,
y formato profesional.

Uso:
    python crear_excel_template.py

Output:
    Template_Propiedades_InmoBot.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def crear_excel_template():
    """Crea el archivo Excel template con todas las validaciones y formato."""

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Propiedades"

    # ============================================
    # DEFINIR ESTILOS
    # ============================================

    # Estilo para encabezados
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Estilo para campos obligatorios
    obligatorio_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================
    # DEFINIR COLUMNAS
    # ============================================

    columnas = [
        # ID (OBLIGATORIO)
        {'letra': 'A', 'nombre': 'id', 'titulo': 'ID (PROP-NNN)', 'ancho': 15, 'obligatorio': True},

        # B√°sicas (OBLIGATORIAS)
        {'letra': 'B', 'nombre': 'tipo', 'titulo': 'Tipo de Propiedad', 'ancho': 18, 'obligatorio': True},
        {'letra': 'C', 'nombre': 'operacion', 'titulo': 'Operaci√≥n', 'ancho': 15, 'obligatorio': True},
        {'letra': 'D', 'nombre': 'estado_construccion', 'titulo': 'Estado Construcci√≥n', 'ancho': 18, 'obligatorio': False},
        {'letra': 'E', 'nombre': 'titulo', 'titulo': 'T√≠tulo (opcional)', 'ancho': 40, 'obligatorio': False},

        # Direcci√≥n (OBLIGATORIAS)
        {'letra': 'F', 'nombre': 'calle', 'titulo': 'Calle y N√∫mero', 'ancho': 30, 'obligatorio': True},
        {'letra': 'G', 'nombre': 'barrio', 'titulo': 'Barrio/Localidad', 'ancho': 20, 'obligatorio': True},
        {'letra': 'H', 'nombre': 'ciudad', 'titulo': 'Ciudad', 'ancho': 15, 'obligatorio': True},
        {'letra': 'I', 'nombre': 'codigo_postal', 'titulo': 'C√≥digo Postal', 'ancho': 12, 'obligatorio': False},

        # Precio (OBLIGATORIAS)
        {'letra': 'J', 'nombre': 'precio_valor', 'titulo': 'Precio', 'ancho': 12, 'obligatorio': True},
        {'letra': 'K', 'nombre': 'precio_moneda', 'titulo': 'Moneda', 'ancho': 10, 'obligatorio': True},
        {'letra': 'L', 'nombre': 'expensas', 'titulo': 'Expensas', 'ancho': 12, 'obligatorio': False},

        # Caracter√≠sticas
        {'letra': 'M', 'nombre': 'ambientes', 'titulo': 'Ambientes', 'ancho': 10, 'obligatorio': False},
        {'letra': 'N', 'nombre': 'dormitorios', 'titulo': 'Dormitorios', 'ancho': 12, 'obligatorio': False},
        {'letra': 'O', 'nombre': 'ba√±os', 'titulo': 'Ba√±os', 'ancho': 10, 'obligatorio': False},
        {'letra': 'P', 'nombre': 'superficie_total', 'titulo': 'Superficie Total', 'ancho': 18, 'obligatorio': False},
        {'letra': 'Q', 'nombre': 'superficie_cubierta', 'titulo': 'Superficie Cubierta', 'ancho': 18, 'obligatorio': False},

        # Checkboxes (S√≠/No)
        {'letra': 'R', 'nombre': 'ascensor', 'titulo': 'Ascensor', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'S', 'nombre': 'balcon', 'titulo': 'Balc√≥n', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'T', 'nombre': 'cochera', 'titulo': 'Cochera', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'U', 'nombre': 'baulera', 'titulo': 'Baulera', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'V', 'nombre': 'patio', 'titulo': 'Patio', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'W', 'nombre': 'pileta', 'titulo': 'Pileta', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'X', 'nombre': 'quincho', 'titulo': 'Quincho', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'Y', 'nombre': 'parrilla', 'titulo': 'Parrilla', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},
        {'letra': 'Z', 'nombre': 'mascotas', 'titulo': 'Mascotas', 'ancho': 10, 'obligatorio': False, 'tipo': 'checkbox'},

        # Otros
        {'letra': 'AA', 'nombre': 'descripcion', 'titulo': 'Descripci√≥n', 'ancho': 50, 'obligatorio': True},
        {'letra': 'AB', 'nombre': 'disponibilidad', 'titulo': 'Disponibilidad', 'ancho': 15, 'obligatorio': False},
    ]

    # ============================================
    # CREAR ENCABEZADOS
    # ============================================

    for col in columnas:
        celda = ws[f"{col['letra']}1"]
        celda.value = col['titulo']
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

        # Ajustar ancho de columna
        ws.column_dimensions[col['letra']].width = col['ancho']

        # Marcar campos obligatorios con color
        if col.get('obligatorio', False):
            celda.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

    # ============================================
    # AGREGAR FILAS DE EJEMPLO CON INSTRUCCIONES
    # ============================================

    # Fila 2: Ejemplo casa usada
    ws['A2'] = 'PROP-001'
    ws['B2'] = 'Casa'
    ws['C2'] = 'Venta'
    ws['D2'] = 'Usado'
    ws['E2'] = 'Casa 3 amb con jard√≠n en Villa Urquiza (opcional, se auto-genera si vac√≠o)'
    ws['F2'] = 'Bauness 2145'
    ws['G2'] = 'Villa Urquiza'
    ws['H2'] = 'CABA'
    ws['I2'] = 'C1431'
    ws['J2'] = 295000
    ws['K2'] = 'USD'
    ws['L2'] = ''
    ws['M2'] = 4
    ws['N2'] = 3
    ws['O2'] = 2
    ws['P2'] = '180 m¬≤'
    ws['Q2'] = '120 m¬≤'
    ws['R2'] = 'No'
    ws['S2'] = 'No'
    ws['T2'] = 'S√≠'
    ws['U2'] = 'No'
    ws['V2'] = 'S√≠'
    ws['W2'] = 'No'
    ws['X2'] = 'S√≠'
    ws['Y2'] = 'S√≠'
    ws['Z2'] = 'S√≠'
    ws['AA2'] = 'Hermosa casa reciclada con jard√≠n y parrilla. Living comedor, cocina integrada, 3 dormitorios.'
    ws['AB2'] = 'Inmediata'

    # Fila 3: Ejemplo departamento a estrenar (en pozo)
    ws['A3'] = 'PROP-002'
    ws['B3'] = 'Departamento'
    ws['C3'] = 'Venta'
    ws['D3'] = 'En pozo'
    ws['E3'] = ''
    ws['F3'] = 'Gorriti 4532, Piso 3¬∞ B'
    ws['G3'] = 'Palermo'
    ws['H3'] = 'CABA'
    ws['I3'] = 'C1414'
    ws['J3'] = 180000
    ws['K3'] = 'USD'
    ws['L3'] = ''
    ws['M3'] = 2
    ws['N3'] = 1
    ws['O3'] = 1
    ws['P3'] = '45 m¬≤'
    ws['Q3'] = '42 m¬≤'
    ws['R3'] = 'S√≠'
    ws['S3'] = 'S√≠'
    ws['T3'] = 'No'
    ws['U3'] = 'No'
    ws['V3'] = 'No'
    ws['W3'] = 'No'
    ws['X3'] = 'No'
    ws['Y3'] = 'No'
    ws['Z3'] = 'No'
    ws['AA3'] = 'Depto a estrenar en pozo de 2 ambientes con balc√≥n. Cocina equipada, ba√±o completo. Entrega Marzo 2026.'
    ws['AB3'] = '2026-03-01'

    # ============================================
    # VALIDACIONES DE DATOS
    # ============================================

    # Validaci√≥n para "Tipo de Propiedad" (columna B)
    dv_tipo = DataValidation(
        type="list",
        formula1='"Departamento,Casa,Local Comercial,Oficina,Lote,Terreno,Campo,Cochera,Galp√≥n"',
        allow_blank=False
    )
    dv_tipo.error = 'Por favor, seleccione un tipo v√°lido de la lista'
    dv_tipo.errorTitle = 'Tipo inv√°lido'
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f'B2:B1000')

    # Validaci√≥n para "Operaci√≥n" (columna C)
    dv_operacion = DataValidation(
        type="list",
        formula1='"Venta,Alquiler,Alquiler Temporario"',
        allow_blank=False
    )
    dv_operacion.error = 'Por favor, seleccione una operaci√≥n v√°lida'
    dv_operacion.errorTitle = 'Operaci√≥n inv√°lida'
    ws.add_data_validation(dv_operacion)
    dv_operacion.add(f'C2:C1000')

    # Validaci√≥n para "Estado Construcci√≥n" (columna D)
    dv_estado_const = DataValidation(
        type="list",
        formula1='"Usado,A estrenar,En pozo,En construcci√≥n,Semi construida,A reciclar"',
        allow_blank=True
    )
    dv_estado_const.error = 'Por favor, seleccione un estado de construcci√≥n v√°lido'
    dv_estado_const.errorTitle = 'Estado construcci√≥n inv√°lido'
    ws.add_data_validation(dv_estado_const)
    dv_estado_const.add(f'D2:D1000')

    # Validaci√≥n para "Moneda"
    dv_moneda = DataValidation(
        type="list",
        formula1='"USD,ARS,EUR"',
        allow_blank=False
    )
    dv_moneda.error = 'Por favor, seleccione una moneda v√°lida'
    dv_moneda.errorTitle = 'Moneda inv√°lida'
    ws.add_data_validation(dv_moneda)
    dv_moneda.add(f'K2:K1000')

    # Validaci√≥n para Checkboxes (S√≠/No)
    dv_sino = DataValidation(
        type="list",
        formula1='"S√≠,No"',
        allow_blank=True
    )
    dv_sino.error = 'Por favor, escriba "S√≠" o "No"'
    dv_sino.errorTitle = 'Valor inv√°lido'
    ws.add_data_validation(dv_sino)
    for col_letra in ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
        dv_sino.add(f'{col_letra}2:{col_letra}1000')

    # Validaci√≥n para n√∫meros (precio, expensas, superficies)
    dv_numero = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1=0,
        allow_blank=True
    )
    dv_numero.error = 'Debe ser un n√∫mero mayor a 0'
    dv_numero.errorTitle = 'N√∫mero inv√°lido'
    ws.add_data_validation(dv_numero)
    for col_letra in ['J', 'L', 'M', 'N', 'O', 'P', 'Q']:
        dv_numero.add(f'{col_letra}2:{col_letra}1000')

    # ============================================
    # CREAR HOJA DE INSTRUCCIONES
    # ============================================

    ws_instrucciones = wb.create_sheet("üìã INSTRUCCIONES")

    instrucciones = [
        ["INSTRUCCIONES PARA COMPLETAR EL TEMPLATE DE PROPIEDADES", ""],
        ["", ""],
        ["1. CAMPOS OBLIGATORIOS (encabezados en ROJO)", ""],
        ["   - ID (formato PROP-NNN, ej: PROP-001, PROP-002)", ""],
        ["   - Tipo de Propiedad, Operaci√≥n", ""],
        ["   - Calle, Barrio, Ciudad", ""],
        ["   - Precio, Moneda", ""],
        ["   - Descripci√≥n", ""],
        ["", ""],
        ["2. CAMPO ID - MUY IMPORTANTE", ""],
        ["   - Formato: PROP-001, PROP-002, PROP-003, etc.", ""],
        ["   - Debe ser √öNICO y SECUENCIAL", ""],
        ["   - Una vez asignado, NO CAMBIAR el ID de una propiedad", ""],
        ["   - Para actualizaciones mensuales: mantener el mismo ID", ""],
        ["   - Propiedades nuevas: usar el siguiente n√∫mero disponible", ""],
        ["   - Propiedades vendidas/alquiladas: ELIMINAR la fila completa", ""],
        ["", ""],
        ["3. ESTADO CONSTRUCCI√ìN (opcional)", ""],
        ["   - Usado (default)", ""],
        ["   - A estrenar", ""],
        ["   - En pozo (departamentos/casas en construcci√≥n preventa)", ""],
        ["   - En construcci√≥n", ""],
        ["   - Semi construida (construcci√≥n parcial)", ""],
        ["   - A reciclar", ""],
        ["", ""],
        ["4. T√çTULO (opcional)", ""],
        ["   - Si lo dejas vac√≠o, se generar√° autom√°ticamente", ""],
        ["   - Formato: 'Tipo Ambientes Operaci√≥n - Barrio'", ""],
        ["   - Ejemplo: 'Casa 3amb Venta - Villa Urquiza'", ""],
        ["", ""],
        ["5. SUPERFICIES (columnas P y Q)", ""],
        ["   - Incluir la unidad: '180 m¬≤' o '7.2 Ha'", ""],
        ["   - Unidades v√°lidas: m¬≤ (metros), Ha (hect√°reas)", ""],
        ["   - Ejemplos: '462 m¬≤', '5.5 Ha', '120 m¬≤'", ""],
        ["", ""],
        ["6. CHECKBOXES (S√≠/No) - Columnas R-Z", ""],
        ["   - Escribir 'S√≠' o 'No' (sin acento tambi√©n funciona)", ""],
        ["   - Dejar vac√≠o = No", ""],
        ["   - Campos disponibles:", ""],
        ["     R: Ascensor  |  S: Balc√≥n    |  T: Cochera  |  U: Baulera", ""],
        ["     V: Patio     |  W: Pileta    |  X: Quincho  |  Y: Parrilla", ""],
        ["     Z: Mascotas", ""],
        ["", ""],
        ["7. FOTOS", ""],
        ["   - Las carpetas deben estar numeradas (1, 2, 3, etc.)", ""],
        ["   - Dentro de cada carpeta: 01.jpg, 02.jpg, 03.jpg, etc.", ""],
        ["   - Estructura esperada:", ""],
        ["     1/", ""],
        ["       ‚îú‚îÄ‚îÄ 01.jpg", ""],
        ["       ‚îú‚îÄ‚îÄ 02.jpg", ""],
        ["       ‚îî‚îÄ‚îÄ 03.jpg", ""],
        ["     2/", ""],
        ["       ‚îú‚îÄ‚îÄ 01.jpg", ""],
        ["       ‚îî‚îÄ‚îÄ 02.jpg", ""],
        ["", ""],
        ["8. PROCESO PARA ACTUALIZACIONES MENSUALES", ""],
        ["   1. Abrir el Excel del mes anterior", ""],
        ["   2. ELIMINAR filas de propiedades vendidas/alquiladas", ""],
        ["   3. MODIFICAR filas de propiedades con cambios (precio, etc.)", ""],
        ["   4. AGREGAR filas nuevas al final con IDs secuenciales", ""],
        ["   5. Guardar el archivo", ""],
        ["   6. Preparar las fotos en carpetas numeradas correspondiendo cada", ""],
        ["      fila del excel con el numero de carpeta de las fotos", ""],
        ["", ""],
        ["9. VALIDACIONES", ""],
        ["   - Los campos con listas desplegables NO permiten otros valores", ""],
        ["   - Los n√∫meros deben ser mayores a 0", ""],
        ["   - Las fechas en formato: AAAA-MM-DD (ej: 2025-02-15)", ""],
        ["", ""],
        ["¬øDudas? Contactar a soporte t√©cnico", ""],
    ]

    for i, fila in enumerate(instrucciones, start=1):
        ws_instrucciones[f'A{i}'] = fila[0]
        if i == 1:
            ws_instrucciones[f'A{i}'].font = Font(size=14, bold=True, color='2563EB')
        elif fila[0].startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.')):
            ws_instrucciones[f'A{i}'].font = Font(size=12, bold=True)

    ws_instrucciones.column_dimensions['A'].width = 80

    # ============================================
    # GUARDAR ARCHIVO
    # ============================================

    filename = f"Template_Propiedades_InmoBot_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    print(f"Excel template creado exitosamente: {filename}")
    print(f"Incluye:")
    print(f"   - {len(columnas)} columnas con validaciones")
    print(f"   - 2 filas de ejemplo")
    print(f"   - Hoja de instrucciones completa")
    print(f"   - Listas desplegables para campos clave")
    print(f"   - Checkboxes para caracteristicas (Si/No)")
    print(f"\nProximo paso: Ejecutar excel_to_json.py con este template completado")

    return filename

if __name__ == '__main__':
    crear_excel_template()
