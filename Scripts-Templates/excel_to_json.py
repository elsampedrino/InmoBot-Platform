#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXCEL TO JSON CONVERTER - INMOBILIARIAS
========================================
Script para convertir el Excel de propiedades de una inmobiliaria al formato JSON
requerido por InmoBot. Procesa el template completado y genera el JSON final.

Soporta actualizaciones mensuales mediante merge con JSON anterior, conservando
las URLs de fotos en Cloudinary para propiedades existentes.

Uso:
    python excel_to_json.py archivo.xlsx [--json-anterior archivo.json]

Parámetros:
    archivo.xlsx      : Archivo Excel con las propiedades
    --json-anterior   : JSON anterior para hacer merge de URLs de fotos (opcional)

Output:
    propiedades_[nombre].json

Ejemplos:
    # Primera vez (sin merge):
    python excel_to_json.py datos_bbr.xlsx
    > Genera: propiedades_bbr.json

    # Actualización mensual (con merge):
    python excel_to_json.py datos_bbr_enero.xlsx --json-anterior propiedades_bbr.json
    > Genera: propiedades_bbr_enero.json conservando URLs de fotos existentes
"""

import sys
import json
import argparse
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def normalizar_texto(texto):
    """Normaliza texto eliminando acentos y caracteres especiales."""
    if not texto:
        return ''

    # Mapeo de caracteres acentuados
    reemplazos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'ñ': 'n', 'Ñ': 'N'
    }

    for acento, sin_acento in reemplazos.items():
        texto = texto.replace(acento, sin_acento)

    return texto

def convertir_bool(valor):
    """Convierte valores Sí/No a boolean."""
    if not valor:
        return False

    valor_str = str(valor).strip().lower()
    return valor_str in ['sí', 'si', 'yes', 'true', '1']

def cargar_json_anterior(archivo_json):
    """
    Carga el JSON anterior y crea un diccionario de propiedades por ID.

    Returns:
        dict: Diccionario {id: propiedad_data} con las URLs de fotos
    """
    if not archivo_json:
        return {}

    try:
        with open(archivo_json, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Crear diccionario indexado por ID
        propiedades_dict = {}
        for prop in data.get('propiedades', []):
            prop_id = prop.get('id')
            if prop_id:
                propiedades_dict[prop_id] = prop

        print(f"JSON anterior cargado: {len(propiedades_dict)} propiedades encontradas")
        return propiedades_dict

    except FileNotFoundError:
        print(f"ADVERTENCIA: Archivo JSON anterior no encontrado: {archivo_json}")
        return {}
    except Exception as e:
        print(f"ERROR al cargar JSON anterior: {e}")
        return {}

def generar_titulo_auto(tipo, ambientes, operacion, barrio):
    """Genera título automáticamente si está vacío."""
    partes = [tipo]

    if ambientes:
        partes.append(f"{ambientes}amb")

    partes.append(operacion)

    if barrio:
        partes.append(f"- {barrio}")

    return " ".join(partes)

def procesar_excel(archivo_excel, json_anterior_dict=None):
    """
    Procesa el archivo Excel y genera el JSON de propiedades.

    Args:
        archivo_excel: Path al archivo Excel
        json_anterior_dict: Diccionario de propiedades anteriores (opcional)
                           Si se provee, conserva URLs de fotos para IDs existentes

    Returns:
        dict: Estructura JSON con las propiedades
    """

    print(f"Procesando archivo: {archivo_excel}")
    if json_anterior_dict:
        print(f"Modo ACTUALIZACIÓN: {len(json_anterior_dict)} propiedades en JSON anterior")

    # Cargar Excel
    try:
        wb = load_workbook(archivo_excel)
        ws = wb['Propiedades']
    except Exception as e:
        print(f"ERROR: No se pudo abrir el archivo Excel: {e}")
        return None

    propiedades = []
    errores = []
    ids_procesados = set()  # Para detectar IDs duplicados

    # Procesar cada fila (desde la 2, la 1 son encabezados)
    fila_num = 2

    while True:
        # Leer ID (columna A)
        propiedad_id = ws[f'A{fila_num}'].value

        # Si no hay ID, asumimos que terminaron las propiedades
        if not propiedad_id:
            break

        try:
            # ============================================
            # LEER DATOS DE LA FILA
            # ============================================

            # ID
            propiedad_id = str(propiedad_id).strip()

            # Básicos
            tipo = ws[f'B{fila_num}'].value
            operacion = ws[f'C{fila_num}'].value
            estado_construccion = ws[f'D{fila_num}'].value or 'Usado'
            titulo = ws[f'E{fila_num}'].value

            # Dirección
            calle = ws[f'F{fila_num}'].value
            barrio = ws[f'G{fila_num}'].value
            ciudad = ws[f'H{fila_num}'].value
            codigo_postal = ws[f'I{fila_num}'].value

            # Precio
            precio_valor = ws[f'J{fila_num}'].value
            precio_moneda = ws[f'K{fila_num}'].value
            expensas = ws[f'L{fila_num}'].value

            # Características
            ambientes = ws[f'M{fila_num}'].value
            dormitorios = ws[f'N{fila_num}'].value
            banios = ws[f'O{fila_num}'].value
            superficie_total = ws[f'P{fila_num}'].value
            superficie_cubierta = ws[f'Q{fila_num}'].value

            # Checkboxes
            ascensor = convertir_bool(ws[f'R{fila_num}'].value)
            balcon = convertir_bool(ws[f'S{fila_num}'].value)
            cochera = convertir_bool(ws[f'T{fila_num}'].value)
            baulera = convertir_bool(ws[f'U{fila_num}'].value)
            patio = convertir_bool(ws[f'V{fila_num}'].value)
            pileta = convertir_bool(ws[f'W{fila_num}'].value)
            quincho = convertir_bool(ws[f'X{fila_num}'].value)
            parrilla = convertir_bool(ws[f'Y{fila_num}'].value)
            mascotas = convertir_bool(ws[f'Z{fila_num}'].value)

            # Otros
            descripcion = ws[f'AA{fila_num}'].value
            disponibilidad = ws[f'AB{fila_num}'].value

            # ============================================
            # VALIDAR CAMPOS OBLIGATORIOS
            # ============================================

            campos_faltantes = []
            if not propiedad_id: campos_faltantes.append('ID')
            if not tipo: campos_faltantes.append('Tipo')
            if not operacion: campos_faltantes.append('Operación')
            if not calle: campos_faltantes.append('Calle')
            if not barrio: campos_faltantes.append('Barrio')
            if not ciudad: campos_faltantes.append('Ciudad')
            if not precio_valor: campos_faltantes.append('Precio')
            if not precio_moneda: campos_faltantes.append('Moneda')
            if not descripcion: campos_faltantes.append('Descripción')

            if campos_faltantes:
                errores.append(f"Fila {fila_num}: Faltan campos obligatorios: {', '.join(campos_faltantes)}")
                fila_num += 1
                continue

            # ============================================
            # VALIDAR ID ÚNICO
            # ============================================

            if propiedad_id in ids_procesados:
                errores.append(f"Fila {fila_num}: ID duplicado: {propiedad_id}")
                fila_num += 1
                continue

            ids_procesados.add(propiedad_id)

            # ============================================
            # GENERAR TÍTULO SI ESTÁ VACÍO
            # ============================================

            if not titulo or titulo.strip() == '':
                titulo = generar_titulo_auto(tipo, ambientes, operacion, barrio)

            # ============================================
            # CONSTRUIR OBJETO PROPIEDAD
            # ============================================

            # Verificar si esta propiedad existe en JSON anterior
            propiedad_anterior = None
            if json_anterior_dict:
                propiedad_anterior = json_anterior_dict.get(propiedad_id)

            # Determinar fotos (conservar URLs si existe, sino placeholder)
            if propiedad_anterior and 'fotos' in propiedad_anterior:
                # Conservar URLs de Cloudinary del JSON anterior
                fotos = propiedad_anterior['fotos']
                estado_fotos = "CONSERVADAS"
            else:
                # Nueva propiedad o sin fotos: usar carpeta numerada como placeholder
                fotos = {
                    "carpeta": str(fila_num - 1),  # fila 2 = carpeta 1
                    "urls": []  # Vacío, se llenará con script de Cloudinary
                }
                estado_fotos = "PENDIENTES"

            # ============================================
            # CONSTRUIR OBJETO OPTIMIZADO (solo campos con datos)
            # ============================================

            propiedad = {
                "id": propiedad_id,
                "tipo": tipo,
                "operacion": operacion,
                "estado_construccion": estado_construccion,
                "titulo": titulo,
                "direccion": {
                    "calle": calle,
                    "barrio": barrio,
                    "ciudad": ciudad
                },
                "precio": {
                    "valor": float(precio_valor) if precio_valor else 0,
                    "moneda": precio_moneda
                },
                "descripcion": descripcion,
                "fotos": fotos
            }

            # Agregar CP solo si existe
            if codigo_postal:
                propiedad["direccion"]["cp"] = codigo_postal

            # Agregar expensas solo si existe
            if expensas:
                propiedad["expensas"] = float(expensas)

            # Construir características solo con valores existentes
            caracteristicas = {}
            if ambientes:
                caracteristicas["ambientes"] = int(ambientes)
            if dormitorios:
                caracteristicas["dormitorios"] = int(dormitorios)
            if banios:
                caracteristicas["banios"] = int(banios)
            if superficie_total:
                caracteristicas["superficie_total"] = str(superficie_total)
            if superficie_cubierta:
                caracteristicas["superficie_cubierta"] = str(superficie_cubierta)

            if caracteristicas:
                propiedad["caracteristicas"] = caracteristicas

            # Construir detalles como ARRAY de strings (solo los que son True)
            detalles = []
            if ascensor:
                detalles.append("ascensor")
            if balcon:
                detalles.append("balcon")
            if cochera:
                detalles.append("cochera")
            if baulera:
                detalles.append("baulera")
            if patio:
                detalles.append("patio")
            if pileta:
                detalles.append("pileta")
            if quincho:
                detalles.append("quincho")
            if parrilla:
                detalles.append("parrilla")
            if mascotas:
                detalles.append("mascotas")

            if detalles:
                propiedad["detalles"] = detalles

            # Agregar disponibilidad solo si es diferente de "Inmediata"
            if disponibilidad and str(disponibilidad) != "Inmediata":
                propiedad["disponibilidad"] = str(disponibilidad)

            propiedades.append(propiedad)
            print(f"  OK Fila {fila_num}: {propiedad_id} - {titulo} [Fotos: {estado_fotos}]")

        except Exception as e:
            errores.append(f"Fila {fila_num}: Error procesando: {str(e)}")

        fila_num += 1

    # ============================================
    # GENERAR RESULTADO
    # ============================================

    print(f"\n{'='*60}")
    print(f"RESUMEN DEL PROCESAMIENTO")
    print(f"{'='*60}")
    print(f"Total propiedades procesadas: {len(propiedades)}")
    print(f"Total errores: {len(errores)}")

    if errores:
        print(f"\nERRORES ENCONTRADOS:")
        for error in errores:
            print(f"  ! {error}")

    # Estructura final del JSON
    resultado = {
        "propiedades": propiedades,
        "metadata": {
            "total": len(propiedades),
            "fecha_generacion": datetime.now().isoformat(),
            "archivo_origen": str(archivo_excel)
        }
    }

    return resultado

def main():
    """Función principal."""

    parser = argparse.ArgumentParser(
        description='Convierte Excel de propiedades a JSON para InmoBot'
    )
    parser.add_argument(
        'archivo',
        help='Archivo Excel con las propiedades'
    )
    parser.add_argument(
        '--json-anterior',
        dest='json_anterior',
        help='JSON anterior para merge de URLs de fotos (opcional)'
    )

    args = parser.parse_args()

    # Validar que el archivo Excel existe
    archivo_path = Path(args.archivo)
    if not archivo_path.exists():
        print(f"ERROR: El archivo {args.archivo} no existe")
        sys.exit(1)

    # Cargar JSON anterior si se especificó
    json_anterior_dict = None
    if args.json_anterior:
        json_anterior_path = Path(args.json_anterior)
        if not json_anterior_path.exists():
            print(f"ADVERTENCIA: JSON anterior no encontrado: {args.json_anterior}")
            print("Continuando sin merge...")
        else:
            json_anterior_dict = cargar_json_anterior(json_anterior_path)

    # Procesar Excel
    resultado = procesar_excel(archivo_path, json_anterior_dict)

    if not resultado:
        print("ERROR: No se pudo procesar el archivo")
        sys.exit(1)

    # Generar nombre de archivo de salida
    nombre_base = archivo_path.stem.lower().replace(' ', '_')
    archivo_salida = f"propiedades_{nombre_base}.json"

    # Guardar JSON
    try:
        with open(archivo_salida, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        print(f"\n{'='*60}")
        print(f"JSON generado exitosamente: {archivo_salida}")
        print(f"{'='*60}")

        # Contar propiedades con fotos pendientes vs conservadas
        fotos_pendientes = sum(1 for p in resultado['propiedades'] if not p['fotos'].get('urls'))
        fotos_conservadas = len(resultado['propiedades']) - fotos_pendientes

        print(f"\nEstado de fotos:")
        print(f"  - Conservadas (URLs existentes): {fotos_conservadas}")
        print(f"  - Pendientes (nuevas/modificadas): {fotos_pendientes}")

        print(f"\nProximos pasos:")
        print(f"1. Verificar el JSON generado")
        if fotos_pendientes > 0:
            print(f"2. Preparar carpetas de fotos numeradas solo para propiedades nuevas/modificadas")
            print(f"3. Ejecutar: python subir_fotos_cloudinary.py {archivo_salida} --carpeta-fotos ./fotos")
        else:
            print(f"2. No hay fotos pendientes - listo para subir a GitHub")

    except Exception as e:
        print(f"ERROR al guardar el archivo JSON: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
