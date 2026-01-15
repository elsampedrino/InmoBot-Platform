#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agregar campo 'Activa' al Excel de BBR
======================================
Script para agregar la columna 'Activa' al Excel existente de Cristian,
rellenar las 34 propiedades con 'Sí', y configurar validaciones para nuevas filas.
"""

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import shutil

def agregar_campo_activa():
    """Agrega el campo 'Activa' al Excel de BBR"""

    input_file = '../BBR Grupo Inmobiliario/BBR_Propiedades_Estandar_20251227.xlsx'

    # Crear backup
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = f'../BBR Grupo Inmobiliario/BBR_Propiedades_Estandar_20251227_backup_{timestamp}.xlsx'
    shutil.copy(input_file, backup_file)
    print(f'✅ Backup creado: {backup_file}')

    # Abrir Excel
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    print(f'\n📊 Excel cargado: {input_file}')
    print(f'   Propiedades: {ws.max_row - 1}')
    print(f'   Columnas actuales: {ws.max_column}')

    # ============================================
    # 1. AGREGAR ENCABEZADO EN COLUMNA AC
    # ============================================

    # Estilos del header (copiar del resto)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')  # Rojo (obligatorio)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Crear encabezado
    ws['AC1'] = 'Activa'
    ws['AC1'].font = header_font
    ws['AC1'].fill = header_fill
    ws['AC1'].alignment = header_alignment
    ws['AC1'].border = thin_border
    ws.column_dimensions['AC'].width = 10

    print(f'\n✅ Encabezado "Activa" agregado en columna AC')

    # ============================================
    # 2. RELLENAR TODAS LAS PROPIEDADES CON "Sí"
    # ============================================

    propiedades_actualizadas = 0
    for fila_num in range(2, ws.max_row + 1):
        # Verificar que la fila tiene datos (ID no vacío)
        id_prop = ws[f'A{fila_num}'].value
        if id_prop:
            ws[f'AC{fila_num}'] = 'Sí'
            propiedades_actualizadas += 1

    print(f'✅ {propiedades_actualizadas} propiedades marcadas como Activa=Sí')

    # ============================================
    # 3. CONFIGURAR VALIDACIONES
    # ============================================

    # Validación Sí/No para columna AC (Activa)
    dv_activa = DataValidation(
        type="list",
        formula1='"Sí,No"',
        allow_blank=False
    )
    dv_activa.error = 'Por favor, escriba "Sí" o "No"'
    dv_activa.errorTitle = 'Valor inválido'
    dv_activa.prompt = 'Indique si la propiedad está activa en el catálogo'
    dv_activa.promptTitle = 'Activa'
    ws.add_data_validation(dv_activa)
    dv_activa.add('AC2:AC1000')

    # Agregar también validaciones a las otras columnas para filas nuevas (36+)
    # Esto ayuda a Cristian cuando agregue propiedades nuevas

    # Tipo de Propiedad (columna B)
    dv_tipo = DataValidation(
        type="list",
        formula1='"Departamento,Casa,Local Comercial,Oficina,Lote,Terreno,Campo,Cochera,Galpón"',
        allow_blank=False
    )
    dv_tipo.error = 'Por favor, seleccione un tipo válido de la lista'
    dv_tipo.errorTitle = 'Tipo inválido'
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f'B36:B1000')  # Solo para filas nuevas

    # Operación (columna C)
    dv_operacion = DataValidation(
        type="list",
        formula1='"Venta,Alquiler,Alquiler Temporario"',
        allow_blank=False
    )
    dv_operacion.error = 'Por favor, seleccione una operación válida'
    dv_operacion.errorTitle = 'Operación inválida'
    ws.add_data_validation(dv_operacion)
    dv_operacion.add(f'C36:C1000')

    # Estado Construcción (columna D)
    dv_estado = DataValidation(
        type="list",
        formula1='"Usado,A estrenar,En pozo,En construcción,Semi construida,A reciclar"',
        allow_blank=True
    )
    dv_estado.error = 'Por favor, seleccione un estado válido'
    dv_estado.errorTitle = 'Estado inválido'
    ws.add_data_validation(dv_estado)
    dv_estado.add(f'D36:D1000')

    # Moneda (columna K)
    dv_moneda = DataValidation(
        type="list",
        formula1='"USD,ARS,EUR"',
        allow_blank=False
    )
    dv_moneda.error = 'Por favor, seleccione una moneda válida'
    dv_moneda.errorTitle = 'Moneda inválida'
    ws.add_data_validation(dv_moneda)
    dv_moneda.add(f'K36:K1000')

    # Checkboxes Sí/No (columnas R-Z y AC)
    dv_sino = DataValidation(
        type="list",
        formula1='"Sí,No"',
        allow_blank=True
    )
    dv_sino.error = 'Por favor, escriba "Sí" o "No"'
    dv_sino.errorTitle = 'Valor inválido'
    ws.add_data_validation(dv_sino)
    for col_letra in ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
        dv_sino.add(f'{col_letra}36:{col_letra}1000')

    print(f'✅ Validaciones configuradas para filas nuevas (36+)')

    # ============================================
    # 4. GUARDAR EXCEL ACTUALIZADO
    # ============================================

    output_file = '../BBR Grupo Inmobiliario/BBR_Propiedades_Estandar_20260102.xlsx'
    wb.save(output_file)

    print(f'\n✅ Excel actualizado guardado: {output_file}')
    print(f'\n📋 Resumen:')
    print(f'   • Columnas: 28 → 29 (agregada "Activa")')
    print(f'   • Propiedades actualizadas: {propiedades_actualizadas}')
    print(f'   • Todas marcadas como Activa=Sí por defecto')
    print(f'   • Validaciones configuradas para filas 36-1000')
    print(f'\n📤 Listo para enviar a Cristian!')

if __name__ == '__main__':
    agregar_campo_activa()
