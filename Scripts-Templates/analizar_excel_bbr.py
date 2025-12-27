#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Script temporal para analizar el Excel de BBR"""

from openpyxl import load_workbook
import sys

# Cargar Excel
wb = load_workbook("../BBR Grupo Inmobiliario/BASE DE DATOS PROPIEDADES BOOT.xlsx")

print("=" * 80)
print("ANÁLISIS COMPLETO DEL EXCEL BBR")
print("=" * 80)
print(f"\nHojas disponibles: {wb.sheetnames}")

# Analizar cada hoja
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    print(f"\n{'='*80}")
    print(f"HOJA: {sheet_name}")
    print(f"{'='*80}")

    # Leer encabezados
    print("\nEncabezados:")
    encabezados = []
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(1, col).value
        if valor:
            encabezados.append((chr(64+col), valor))
            print(f"  {chr(64+col)}: {valor}")

    print(f"\nTotal filas con datos: {ws.max_row - 1}")

    # Mostrar primeras 2 filas de ejemplo
    if ws.max_row > 1:
        print("\nEJEMPLO DE DATOS (primeras 2 filas):")
        print("-" * 80)

        for fila in range(2, min(4, ws.max_row + 1)):
            print(f"\nFila {fila}:")
            tiene_datos = False
            for col in range(1, ws.max_column + 1):
                celda = ws.cell(fila, col)
                if celda.value:
                    tiene_datos = True
                    letra = chr(64 + col)
                    print(f"  {letra}: {celda.value}")

            if not tiene_datos:
                print("  (vacía)")

print(f"\n{'='*80}")
print("RESUMEN")
print(f"{'='*80}")
print(f"Total hojas: {len(wb.sheetnames)}")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    filas = ws.max_row - 1
    print(f"  - {sheet_name}: {filas} propiedades")
