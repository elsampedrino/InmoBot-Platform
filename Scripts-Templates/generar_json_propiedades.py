"""
Genera propiedades_bbr.json a partir del Template_Propiedades_InmoBot_AAAAMMDD.xlsx
Estructura de salida compatible con propiedades_bbr.json (formato InmoBot v2)

Uso:
    python generar_json_propiedades.py
    python generar_json_propiedades.py Template_Propiedades_InmoBot_20260303.xlsx
"""

import sys
import io
import json
import re
import glob
import os
from pathlib import Path
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("ERROR: instalar openpyxl ->  pip install openpyxl")
    sys.exit(1)

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
MAX_DESTACADAS = 6
OUTPUT_JSON    = "propiedades_bbr.json"

# Columnas del template (índice 0-based dentro de cada fila)
COL = {
    "id":                   0,
    "tipo":                 1,
    "operacion":            2,
    "titulo":               3,
    "destacado":            4,
    "activo":               5,
    "calle":                6,
    "barrio":               7,
    "ciudad":               8,
    "lat":                  9,
    "lng":                 10,
    "precio":              11,
    "moneda":              12,
    "expensas":            13,
    "antiguedad":          14,
    "estado_construccion": 15,
    "ambientes":           16,
    "dormitorios":         17,
    "banios":              18,
    "superficie_total":    19,
    "superficie_cubierta": 20,
    "detalles":            21,
    "descripcion":         22,
    "descripcion_corta":   23,
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def get(row, campo):
    """Lee un valor del row por nombre de campo, retorna None si está fuera de rango."""
    idx = COL[campo]
    return row[idx] if idx < len(row) else None

def limpio(valor):
    """Convierte a string limpio, retorna None si vacío."""
    if valor is None:
        return None
    s = str(valor).strip()
    return s if s else None

def a_bool(valor, default=False):
    """Convierte Sí/Si/si/YES/True → True, resto → False."""
    if valor is None:
        return default
    s = str(valor).strip().lower()
    return s in ("sí", "si", "yes", "true", "1", "verdadero")

def a_numero(valor):
    """Convierte a int si es posible, sino None."""
    if valor is None:
        return None
    try:
        return int(float(str(valor).replace(",", ".")))
    except:
        return None

def a_float(valor):
    """Convierte a float si es posible, sino None."""
    if valor is None:
        return None
    try:
        return float(str(valor).replace(",", "."))
    except:
        return None

def normalizar_tipo(valor):
    """Casa → casa, Departamento → departamento, etc."""
    if not valor:
        return None
    mapa = {
        "casa":            "casa",
        "departamento":    "departamento",
        "depto":           "departamento",
        "lote":            "lote",
        "terreno":         "terreno",
        "campo":           "campo",
        "local comercial": "local_comercial",
        "oficina":         "oficina",
        "cochera":         "cochera",
        "galpon":          "galpon",
        "galpón":          "galpon",
    }
    return mapa.get(str(valor).strip().lower(), str(valor).strip().lower())

def normalizar_operacion(valor):
    """Venta → venta, Alquiler → alquiler, Alquiler Temporario → alquiler_temporario."""
    if not valor:
        return None
    mapa = {
        "venta":               "venta",
        "alquiler":            "alquiler",
        "alquiler temporario": "alquiler_temporario",
    }
    return mapa.get(str(valor).strip().lower(), str(valor).strip().lower())

def normalizar_estado(valor):
    """Usado → usado, A estrenar → a_estrenar, etc."""
    if not valor:
        return None
    mapa = {
        "usado":           "usado",
        "a estrenar":      "a_estrenar",
        "en pozo":         "en_pozo",
        "en construccion": "en_construccion",
        "en construcción": "en_construccion",
        "semi construida": "semi_construida",
        "a reciclar":      "a_reciclar",
    }
    return mapa.get(str(valor).strip().lower(), str(valor).strip().lower())

def normalizar_superficie(valor):
    """
    Acepta número (ej: 180) o string (ej: '180 m²', '5.5 Ha').
    Retorna string con unidad o None.
    """
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    # Si ya tiene unidad, devolver limpio
    if "m²" in s or "m2" in s.lower() or "ha" in s.lower():
        return s.replace("m2", "m²").replace("M2", "m²")
    # Si es puro número, asumir m²
    try:
        n = int(float(s.replace(",", ".")))
        return f"{n} m²"
    except:
        return s

def generar_titulo(tipo, operacion, ciudad, ambientes=None):
    """Genera título automático si no se proveyó."""
    tipo_str = str(tipo).capitalize() if tipo else "Propiedad"
    op_str   = str(operacion).capitalize() if operacion else ""
    amb_str  = f" {ambientes} amb" if ambientes else ""
    ciudad_str = str(ciudad).strip() if ciudad else ""
    return f"{tipo_str}{amb_str} {op_str} - {ciudad_str}".strip(" -")

def parsear_detalles(valor):
    """'patio, pileta, quincho' → ['patio', 'pileta', 'quincho']"""
    if not valor:
        return []
    items = re.split(r"[,;|]", str(valor))
    return [i.strip().lower() for i in items if i.strip()]

def buscar_excel():
    """Busca el template más reciente en el directorio actual."""
    archivos = sorted(glob.glob("Template_Propiedades_InmoBot_*.xlsx"), reverse=True)
    if archivos:
        return archivos[0]
    archivos = sorted(glob.glob("*.xlsx"), reverse=True)
    return archivos[0] if archivos else None

# ─── PROCESAMIENTO PRINCIPAL ─────────────────────────────────────────────────

def procesar(excel_path):
    print(f"Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if "Propiedades" not in wb.sheetnames:
        print(f"ERROR: No se encontró la hoja 'Propiedades'. Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["Propiedades"]
    propiedades = []
    prop_index  = 1  # para auto-asignar carpeta de fotos

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas completamente vacías
        if not any(v is not None for v in row):
            continue

        # ── Identificación ──────────────────────────────────────────────────
        prop_id   = limpio(get(row, "id")) or f"PROP-{str(prop_index).zfill(3)}"
        tipo      = normalizar_tipo(get(row, "tipo"))
        operacion = normalizar_operacion(get(row, "operacion"))
        activo    = a_bool(get(row, "activo"), default=True)
        destacado = a_bool(get(row, "destacado"), default=False)

        if not tipo or not operacion:
            print(f"  FILA {row_idx}: sin tipo/operación, se omite.")
            continue

        if not activo:
            print(f"  FILA {row_idx} ({prop_id}): activo=No, se omite del JSON.")
            prop_index += 1
            continue

        # ── Ubicación ────────────────────────────────────────────────────────
        calle  = limpio(get(row, "calle"))  or ""
        barrio = limpio(get(row, "barrio")) or ""
        ciudad = limpio(get(row, "ciudad")) or ""
        lat    = a_float(get(row, "lat"))
        lng    = a_float(get(row, "lng"))

        # ── Título ───────────────────────────────────────────────────────────
        titulo_raw = limpio(get(row, "titulo"))
        ambientes  = a_numero(get(row, "ambientes"))
        titulo = titulo_raw or generar_titulo(tipo, operacion, ciudad, ambientes)

        # ── Precio ───────────────────────────────────────────────────────────
        precio_val = a_numero(get(row, "precio")) or 0
        moneda     = limpio(get(row, "moneda")) or "USD"
        expensas   = a_numero(get(row, "expensas"))

        # ── Características ──────────────────────────────────────────────────
        antiguedad        = a_numero(get(row, "antiguedad"))   # años como número o None
        estado            = normalizar_estado(get(row, "estado_construccion"))
        dormitorios       = a_numero(get(row, "dormitorios"))
        banios            = a_numero(get(row, "banios"))
        sup_total         = normalizar_superficie(get(row, "superficie_total"))
        sup_cubierta      = normalizar_superficie(get(row, "superficie_cubierta"))

        # ── Descripciones ─────────────────────────────────────────────────────
        descripcion       = limpio(get(row, "descripcion"))       or titulo
        descripcion_corta = limpio(get(row, "descripcion_corta")) or ""

        # ── Detalles ──────────────────────────────────────────────────────────
        detalles = parsear_detalles(get(row, "detalles"))

        # ── Construir objeto ──────────────────────────────────────────────────
        direccion = {"calle": calle, "barrio": barrio, "ciudad": ciudad}
        if lat is not None:
            direccion["lat"] = lat
        if lng is not None:
            direccion["lng"] = lng

        precio = {"valor": precio_val, "moneda": moneda, "expensas": expensas}

        caracteristicas = {
            "antigüedad":         antiguedad,
            "estado_construccion": estado,
            "ambientes":          ambientes,
            "dormitorios":        dormitorios,
            "banios":             banios,
            "superficie_total":   sup_total,
            "superficie_cubierta": sup_cubierta,
        }

        propiedad = {
            "id":               prop_id,
            "tipo":             tipo,
            "operacion":        operacion,
            "titulo":           titulo,
            "destacado":        destacado,
            "activo":           activo,
            "direccion":        direccion,
            "precio":           precio,
            "descripcion":      descripcion,
            "descripcion_corta": descripcion_corta,
            "fotos": {
                "carpeta": str(prop_index),
                "urls":    []
            },
            "caracteristicas":  caracteristicas,
            "detalles":         detalles,
        }

        propiedades.append(propiedad)
        estado_str = "DESTACADA" if destacado else ("ACTIVA" if activo else "INACTIVA")
        print(f"  PROP {prop_id}: {tipo} {operacion} | {calle}, {ciudad} [{estado_str}]")
        prop_index += 1

    wb.close()

    # ── Validación destacadas ─────────────────────────────────────────────────
    destacadas = [p for p in propiedades if p["destacado"]]
    if len(destacadas) > MAX_DESTACADAS:
        print(f"\nADVERTENCIA: {len(destacadas)} propiedades con destacado=true "
              f"(máximo permitido: {MAX_DESTACADAS}).")
        print("  Se conservan las primeras 6 como destacadas, el resto pasa a false.")
        ids_destacadas = {p["id"] for p in destacadas[:MAX_DESTACADAS]}
        for p in propiedades:
            if p["destacado"] and p["id"] not in ids_destacadas:
                p["destacado"] = False
                print(f"  -> {p['id']} cambió a destacado=false")

    # ── Generar JSON ──────────────────────────────────────────────────────────
    output = {
        "metadata": {
            "total":            len(propiedades),
            "fecha_generacion": date.today().isoformat(),
            "fuente":           Path(excel_path).name,
        },
        "propiedades": propiedades,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nJSON generado: {OUTPUT_JSON}")
    print(f"Total propiedades: {len(propiedades)}")

    # Resumen por tipo
    from collections import Counter
    resumen = Counter(f"{p['tipo']} / {p['operacion']}" for p in propiedades)
    for k, v in sorted(resumen.items()):
        print(f"  {v:>3}x  {k}")

    activas    = sum(1 for p in propiedades if p["activo"])
    destacadas = sum(1 for p in propiedades if p["destacado"])
    print(f"\n  Activas:    {activas}")
    print(f"  Destacadas: {destacadas}")


# ─── EJECUCIÓN ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = buscar_excel()
        if not excel_path:
            print("ERROR: No se encontró ningún archivo .xlsx en el directorio actual.")
            print("Uso: python generar_json_propiedades.py Template_Propiedades_InmoBot_AAAAMMDD.xlsx")
            sys.exit(1)
        print(f"Usando template: {excel_path}")

    if not os.path.exists(excel_path):
        print(f"ERROR: No existe el archivo '{excel_path}'")
        sys.exit(1)

    print("=" * 60)
    print("  InmoBot - Generador de JSON de Propiedades")
    print("=" * 60)
    procesar(excel_path)
    print("=" * 60)
    print("Proceso completado.")
    print("\nProximos pasos:")
    print("  1. Completar fotos.urls en el JSON (URLs de Cloudinary)")
    print("  2. Agregar lat/lng a propiedades con mapa")
    print(f"  3. Subir {OUTPUT_JSON} al repositorio GitHub de datos")
