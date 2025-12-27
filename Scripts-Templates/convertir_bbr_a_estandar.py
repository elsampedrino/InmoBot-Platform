#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CONVERTIR EXCEL BBR A FORMATO ESTÁNDAR
========================================
Convierte el Excel multi-hoja de BBR al formato estándar de InmoBot (25 columnas).

Uso:
    python convertir_bbr_a_estandar.py

Input:
    ../BBR Grupo Inmobiliario/BASE DE DATOS PROPIEDADES BOOT.xlsx

Output:
    ../BBR Grupo Inmobiliario/BBR_Propiedades_Estandar_YYYYMMDD.xlsx

Mapeo de columnas BBR -> Estándar:
    A: ID                  -> A: ID (conservar)
    B: Operación           -> C: Operación
    C: Tipo                -> B: Tipo de Propiedad
    E: Dirección           -> F: Calle y Número
    F: Localidad           -> G: Barrio/Localidad
    G: m2 Construidos      -> P: Superficie Cubierta
    H: Descripción         -> X: Descripción
    I: Valor               -> J: Precio

    Campos adicionales:
    - H: Ciudad = "Ramallo" o "Pergamino" (según localidad)
    - K: Moneda = "USD" o "ARS" (detectar desde valor)
    - D: Foto (número de carpeta) -> conservar para fotos
    - D: Estado Construcción = vacío (se puede inferir)
    - E: Título = auto-generar
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re

# Colores
COLOR_OBLIGATORIO = "FFE6E6"  # Rojo claro
COLOR_OPCIONAL = "E6F2FF"     # Azul claro
COLOR_HEADER = "4472C4"       # Azul

def detectar_moneda(valor_str):
    """
    Detecta la moneda desde el string de valor.

    Returns:
        tuple: (moneda, valor_numerico)
    """
    if not valor_str:
        return ("USD", None)

    valor_str = str(valor_str).strip()

    # Patrones comunes
    if valor_str.lower().startswith("u$s"):
        # "u$s 139.000" -> USD, 139000
        numero = re.sub(r'[^\d]', '', valor_str)
        return ("USD", float(numero) if numero else None)

    elif "$" in valor_str.lower() or "usd" in valor_str.lower():
        numero = re.sub(r'[^\d]', '', valor_str)
        return ("USD", float(numero) if numero else None)

    else:
        # Asumir ARS si solo hay números
        numero = re.sub(r'[^\d]', '', valor_str)
        return ("ARS", float(numero) if numero else None)

def limpiar_superficie(sup_str):
    """
    Extrae números de strings de superficie.
    "140" -> 140.0
    "11 x 42" -> 462.0 (área)
    """
    if not sup_str:
        return None

    sup_str = str(sup_str).strip()

    # Si tiene "x" (ej: "11 x 42"), calcular área
    if 'x' in sup_str.lower():
        partes = re.findall(r'\d+(?:\.\d+)?', sup_str)
        if len(partes) >= 2:
            return float(partes[0]) * float(partes[1])

    # Extraer primer número
    numeros = re.findall(r'\d+(?:\.\d+)?', sup_str)
    if numeros:
        return float(numeros[0])

    return None

def normalizar_tipo_propiedad(tipo_bbr):
    """
    Normaliza el tipo de propiedad al formato estándar.
    """
    mapeo = {
        'Casa': 'Casa',
        'Departamento': 'Departamento',
        'Lote': 'Terreno',
        'Campo': 'Campo',
        'Local': 'Local Comercial'
    }
    return mapeo.get(tipo_bbr, tipo_bbr)

def normalizar_operacion(op_bbr):
    """
    Normaliza la operación al formato estándar.
    """
    mapeo = {
        'Venta': 'Venta',
        'Alquiler': 'Alquiler'
    }
    return mapeo.get(op_bbr, op_bbr)

def determinar_ciudad(localidad):
    """
    Determina la ciudad principal desde la localidad.
    """
    if not localidad:
        return "Ramallo"

    localidad_lower = localidad.lower()

    if 'pergamino' in localidad_lower:
        return "Pergamino"
    elif 'villa ramallo' in localidad_lower:
        return "Villa Ramallo"
    else:
        return "Ramallo"

def generar_titulo_auto(tipo, operacion, localidad):
    """
    Genera título automático estilo: "Casa Venta - Ramallo"
    """
    return f"{tipo} {operacion} - {localidad}"

def crear_excel_estandar():
    """
    Crea un nuevo Excel con el formato estándar (25 columnas).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Propiedades"

    # Definir columnas estándar
    columnas = [
        # ID (OBLIGATORIO)
        {'letra': 'A', 'nombre': 'id', 'titulo': 'ID (PROP-NNN)', 'ancho': 15, 'obligatorio': True},

        # Básicas (OBLIGATORIAS)
        {'letra': 'B', 'nombre': 'tipo', 'titulo': 'Tipo de Propiedad', 'ancho': 18, 'obligatorio': True},
        {'letra': 'C', 'nombre': 'operacion', 'titulo': 'Operación', 'ancho': 15, 'obligatorio': True},

        # Opcionales
        {'letra': 'D', 'nombre': 'estado_construccion', 'titulo': 'Estado Construcción', 'ancho': 20, 'obligatorio': False},
        {'letra': 'E', 'nombre': 'titulo', 'titulo': 'Título (opcional)', 'ancho': 35, 'obligatorio': False},

        # Ubicación (OBLIGATORIAS)
        {'letra': 'F', 'nombre': 'calle', 'titulo': 'Calle y Número', 'ancho': 25, 'obligatorio': True},
        {'letra': 'G', 'nombre': 'barrio', 'titulo': 'Barrio/Localidad', 'ancho': 20, 'obligatorio': True},
        {'letra': 'H', 'nombre': 'ciudad', 'titulo': 'Ciudad', 'ancho': 18, 'obligatorio': True},
        {'letra': 'I', 'nombre': 'cp', 'titulo': 'Código Postal', 'ancho': 15, 'obligatorio': False},

        # Precio (OBLIGATORIO)
        {'letra': 'J', 'nombre': 'precio', 'titulo': 'Precio', 'ancho': 15, 'obligatorio': True},
        {'letra': 'K', 'nombre': 'moneda', 'titulo': 'Moneda', 'ancho': 12, 'obligatorio': True},
        {'letra': 'L', 'nombre': 'expensas', 'titulo': 'Expensas', 'ancho': 12, 'obligatorio': False},

        # Características (OPCIONALES)
        {'letra': 'M', 'nombre': 'ambientes', 'titulo': 'Ambientes', 'ancho': 12, 'obligatorio': False},
        {'letra': 'N', 'nombre': 'dormitorios', 'titulo': 'Dormitorios', 'ancho': 12, 'obligatorio': False},
        {'letra': 'O', 'nombre': 'banios', 'titulo': 'Baños', 'ancho': 12, 'obligatorio': False},
        {'letra': 'P', 'nombre': 'sup_total', 'titulo': 'Superficie Total (m²)', 'ancho': 22, 'obligatorio': False},
        {'letra': 'Q', 'nombre': 'sup_cubierta', 'titulo': 'Superficie Cubierta (m²)', 'ancho': 24, 'obligatorio': False},

        # Checkboxes (OPCIONALES)
        {'letra': 'R', 'nombre': 'ascensor', 'titulo': 'Ascensor (SI/NO)', 'ancho': 18, 'obligatorio': False},
        {'letra': 'S', 'nombre': 'balcon', 'titulo': 'Balcón (SI/NO)', 'ancho': 16, 'obligatorio': False},
        {'letra': 'T', 'nombre': 'cochera', 'titulo': 'Cochera (SI/NO)', 'ancho': 17, 'obligatorio': False},
        {'letra': 'U', 'nombre': 'baulera', 'titulo': 'Baulera (SI/NO)', 'ancho': 17, 'obligatorio': False},
        {'letra': 'V', 'nombre': 'pileta', 'titulo': 'Pileta (SI/NO)', 'ancho': 16, 'obligatorio': False},
        {'letra': 'W', 'nombre': 'mascotas', 'titulo': 'Mascotas (SI/NO)', 'ancho': 18, 'obligatorio': False},

        # Descripción (OBLIGATORIA)
        {'letra': 'X', 'nombre': 'descripcion', 'titulo': 'Descripción', 'ancho': 50, 'obligatorio': True},

        # Disponibilidad (OPCIONAL)
        {'letra': 'Y', 'nombre': 'disponibilidad', 'titulo': 'Disponibilidad', 'ancho': 15, 'obligatorio': False},
    ]

    # Crear encabezados
    for col in columnas:
        celda = ws[f"{col['letra']}1"]
        celda.value = col['titulo']
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Ajustar ancho
        ws.column_dimensions[col['letra']].width = col['ancho']

    return wb, ws, columnas

def procesar_hoja_bbr(ws_bbr, sheet_name, ws_destino, fila_actual):
    """
    Procesa una hoja del Excel BBR y escribe en el Excel estándar.

    Returns:
        int: Próxima fila disponible en el Excel destino
    """
    print(f"\nProcesando hoja: {sheet_name}")

    # Leer todas las filas (saltar encabezado)
    propiedades_procesadas = 0

    for fila_bbr in range(2, ws_bbr.max_row + 1):
        # Leer datos BBR
        id_prop = ws_bbr.cell(fila_bbr, 1).value  # A: ID

        if not id_prop:
            continue  # Saltar filas vacías

        operacion_bbr = ws_bbr.cell(fila_bbr, 2).value  # B: Operación
        tipo_bbr = ws_bbr.cell(fila_bbr, 3).value       # C: Tipo
        foto_num = ws_bbr.cell(fila_bbr, 4).value       # D: Foto
        direccion = ws_bbr.cell(fila_bbr, 5).value      # E: Dirección
        localidad = ws_bbr.cell(fila_bbr, 6).value      # F: Localidad

        # Columnas variables según tipo de propiedad
        if sheet_name == "Casas":
            m2_construidos = ws_bbr.cell(fila_bbr, 7).value  # G
            lote = ws_bbr.cell(fila_bbr, 8).value            # H
            descripcion = ws_bbr.cell(fila_bbr, 9).value     # I
            valor = ws_bbr.cell(fila_bbr, 10).value          # J
        elif sheet_name == "Campos":
            hectareas = ws_bbr.cell(fila_bbr, 7).value       # G: Hectáreas
            descripcion = ws_bbr.cell(fila_bbr, 8).value     # H
            valor_ha = ws_bbr.cell(fila_bbr, 9).value        # I: Valor Hectárea
            valor = ws_bbr.cell(fila_bbr, 10).value          # J: Valor Total
            m2_construidos = None
            lote = f"{hectareas} Ha" if hectareas else None
        else:  # Departamentos, Lotes, Locales
            m2_construidos = ws_bbr.cell(fila_bbr, 7).value  # G
            descripcion = ws_bbr.cell(fila_bbr, 8).value     # H
            valor = ws_bbr.cell(fila_bbr, 9).value           # I
            lote = None

        # Procesar datos
        tipo = normalizar_tipo_propiedad(tipo_bbr)
        operacion = normalizar_operacion(operacion_bbr)
        ciudad = determinar_ciudad(localidad)
        titulo = generar_titulo_auto(tipo, operacion, localidad or ciudad)
        moneda, precio = detectar_moneda(valor)
        superficie = limpiar_superficie(m2_construidos)

        # Escribir en Excel estándar
        ws_destino.cell(fila_actual, 1).value = id_prop                    # A: ID
        ws_destino.cell(fila_actual, 2).value = tipo                       # B: Tipo
        ws_destino.cell(fila_actual, 3).value = operacion                  # C: Operación
        # D: Estado Construcción (vacío por ahora)
        ws_destino.cell(fila_actual, 5).value = titulo                     # E: Título
        ws_destino.cell(fila_actual, 6).value = direccion                  # F: Calle
        ws_destino.cell(fila_actual, 7).value = localidad or ciudad        # G: Barrio/Localidad
        ws_destino.cell(fila_actual, 8).value = ciudad                     # H: Ciudad
        # I: CP (vacío)
        ws_destino.cell(fila_actual, 10).value = precio                    # J: Precio
        ws_destino.cell(fila_actual, 11).value = moneda                    # K: Moneda
        # L: Expensas (vacío)
        # M-O: Ambientes, Dormitorios, Baños (vacío)
        ws_destino.cell(fila_actual, 16).value = superficie                # P: Superficie Total
        ws_destino.cell(fila_actual, 17).value = superficie                # Q: Superficie Cubierta
        # R-W: Checkboxes (vacío)
        ws_destino.cell(fila_actual, 24).value = descripcion               # X: Descripción
        ws_destino.cell(fila_actual, 25).value = "Inmediata"               # Y: Disponibilidad

        propiedades_procesadas += 1
        fila_actual += 1

    print(f"  OK Procesadas: {propiedades_procesadas} propiedades")

    return fila_actual

def main():
    """Función principal."""

    print("="*80)
    print("CONVERSIÓN BBR A FORMATO ESTÁNDAR")
    print("="*80)

    # Cargar Excel BBR
    print("\nCargando Excel BBR...")
    try:
        wb_bbr = load_workbook("../BBR Grupo Inmobiliario/BASE DE DATOS PROPIEDADES BOOT.xlsx")
        print(f"OK Excel cargado: {len(wb_bbr.sheetnames)} hojas")
    except Exception as e:
        print(f"ERROR al cargar Excel BBR: {e}")
        return

    # Crear Excel estándar
    print("\nCreando Excel estándar...")
    wb_estandar, ws_estandar, columnas = crear_excel_estandar()
    print(f"OK Excel estándar creado: {len(columnas)} columnas")

    # Procesar cada hoja
    fila_actual = 2  # Empezar después del encabezado

    for sheet_name in wb_bbr.sheetnames:
        ws_bbr = wb_bbr[sheet_name]
        fila_actual = procesar_hoja_bbr(ws_bbr, sheet_name, ws_estandar, fila_actual)

    # Guardar Excel estándar
    fecha = datetime.now().strftime("%Y%m%d")
    archivo_salida = f"../BBR Grupo Inmobiliario/BBR_Propiedades_Estandar_{fecha}.xlsx"

    try:
        wb_estandar.save(archivo_salida)
        print(f"\n{'='*80}")
        print(f"OK CONVERSION EXITOSA")
        print(f"{'='*80}")
        print(f"\nArchivo generado: {archivo_salida}")
        print(f"Total propiedades convertidas: {fila_actual - 2}")
        print(f"\nProximos pasos:")
        print(f"1. Revisar el Excel generado")
        print(f"2. Completar campos faltantes si es necesario (ambientes, banios, etc.)")
        print(f"3. Convertir a JSON con:")
        print(f'   python excel_to_json.py "{archivo_salida}"')
    except Exception as e:
        print(f"ERROR al guardar Excel: {e}")

if __name__ == '__main__':
    main()
